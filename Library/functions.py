from datetime import date
from datetime import datetime
import pandas as pd
from Library.log import writetoerrorlog
import openpyxl
from openpyxl import load_workbook
from itertools import chain
import xlwings as xw


def order():  # Allows customer to order pizza with desired toppings
    total = 0  # Initializes total as $0
    toppings_receipt = []  # Initializes list of toppings for receipt

    pd.options.display.float_format = '{:.2f}'.format  # Forces prices to be displayed with two decimal places

    try:
        base = pd.read_csv('Prices/Base.csv')  # Opens the files with pricing info
        toppings = pd.read_csv('Prices/Topping.csv')
        base_diction = base.to_dict(orient='records')  # Converts files into lists of dictionaries
        toppings_diction = toppings.to_dict(orient='records')
    except Exception as e:
        print("Problem opening base and topping price files")
        writetoerrorlog("order", e)

    name = input("Name: ")  # Requests users name
    number = input("Phone Number: ")  # Requests users phone number
    print(base)  # Displays base size menu

    ## Determines users choice and adds price to total ##

    while True:
        try:
            size = input("Input number associated with desired size.")
            if 0 <= int(size) < 3:
                size_int = int(size)
                size_price = int(base_diction[size_int].get("Price"))  # Determines price from designated dictionary
                total += size_price
                break
            else:
                print("Error: Please input a valid number.")
        except Exception as e:
            print("Error: Please input a valid number.")
            writetoerrorlog("order", e)

    ## Displays toppings menu and requests users preference ##
    print(toppings)

    ## Requests users topping preferences until they have all their desired toppings on the order ##
    while True:
        try:
            top = input("Input number associated with desired topping and type done when finished.")
            if top.upper() == "DONE":  # Continues if user is still choosing toppings
                break
            elif 0 <= int(top) < 13:
                toppings_receipt.append(top)  # Stores selected toppings to list to add to receipt later
                top_int = int(top)
                top_price = int(toppings_diction[top_int].get("Price"))  # Finds topping price based on dictionary
                total += top_price  # Adds topping price to final total
            else:
                print("Error: Please input a valid number.")
                continue
        except Exception as e:
            print("Error: Please input a valid number.")
            writetoerrorlog("order", e)

    ## Creates receipt for pizza order with title in format name_MM-DD-YYYY_HH.MM.SS.txt ##
    try:
        filename = f"{name}_{date.today().strftime('%d-%m-%Y')}_{datetime.now().strftime('%H.%M.%S')}.txt"
        with open(f'Receipts/{filename}', "w") as file:
            r_total = "{:.2f}".format(total)  # Ensures total has two decimal places
            s_price = "{:.2f}".format(size_price)  # Ensures base price has two decimal places
            size_name = base_diction[size_int].get("Size")  # Pulls chosen size from designated dictionary
            file.write(f"Customer: {name} Contact: {number}\n")  # Displays customers name and number on receipt
            file.write(" \n")
            file.write(f"{size_name} - ${s_price}\n")  # Displays customers selected size
            for x in toppings_receipt:  # Displays all toppings selected on receipt
                topping_name = toppings_diction[int(x)].get("Topping")
                topping_price = "{:.2f}".format(toppings_diction[int(x)].get("Price"))
                file.write(f"{topping_name} - ${topping_price}\n")
            file.write(" \n")
            file.write(f"Total - ${r_total}\n")  # Displays total price of pizza order
        top_inventory = []
        for n in toppings_receipt:
            top_inv = toppings_diction[int(n)].get("Topping")
            top_inventory.append(top_inv)
        return name, total, top_inventory, size_name
    except Exception as e:
        print("Error: Please input a valid number.")
        writetoerrorlog("order", e)  #


def subtract_inventory(topping, size):  # Deducts ordered products from inventory
    try:
        # Opens inventory workbook and ensures formatting and formulas stay intact #
        i_wb = load_workbook("pizza.xlsx", data_only=False, keep_links=True)
        i_sh = i_wb["Inventory"]
        p = 17
        while i_sh["A" + str(p)].value != size:  # Finds the size ordered
            p += 1
        size_val = int(i_sh["B" + str(p)].value)
        i_sh["B" + str(p)].value = size_val - 1  # Subtracts one from inventory
        for x in topping:  # Finds all selected toppings and subtracts one
            n = 1  # from inventory
            while i_sh["A" + str(n)].value != x:
                n += 1
            top_val = int(i_sh["B" + str(n)].value)
            i_sh["B" + str(n)].value = top_val - 1
        for x in chain(range(2, 15), range(17, 20)):  # Checks if stock is empty
            level = i_sh["B" + str(x)].value
            if level == 0:
                print("**Stock of " + str(i_sh["A" + str(x)].value) + " is empty**")
            elif 0 < level <= 3:
                print("**Stock of " + str(i_sh["A" + str(x)].value) + " is running low**")
            else:
                continue
        i_wb.save("pizza.xlsx")  # Saves and closes inventory
        i_wb.close()
    except Exception as e:
        print("Error occurred updating inventory.")
        writetoerrorlog("subtract_inventory", e)  #


def sales(name, total):  # Inputs order totals into P and L statements
    try:
        # Opens profit and loss workbook and ensures format and formulas stay intact #
        profit_wb = load_workbook("pizza.xlsx", data_only=False, keep_links=True)
        profit_sh = profit_wb["P and L"]
        n = 2
        while True:  # Finds closest blank box and
            if profit_sh["B" + str(n)].value is None:  # inputs the total and customer
                profit_sh["A" + str(n)].value = name
                profit_sh["B" + str(n)].value = total
                profit_sh["B" + str(n)].number_format = "$#,##0.00"  # Keeps money format in excel
                profit_wb.save("pizza.xlsx")  # Saves and closes wb
                profit_wb.close()
                break
            else:
                n += 1
                continue
    except Exception as e:
        print("Error occurred updating P and L statements.")
        writetoerrorlog("sales", e)


def view_inventory():  # Displays inventory
    try:
        # Opens inventory and ensures everything is intact #
        i_wb = load_workbook("pizza.xlsx", data_only=False, keep_links=True)
        i_sh = i_wb["Inventory"]
        print(" ")
        print("Toppings")
        for x in range(2, 15):  # Displays stock of all toppings
            name = str(i_sh["A" + str(x)].value)
            stock = str(i_sh["B" + str(x)].value)
            print(name + "=> " + stock + " in stock")
        print(" ")
        print("Bases")
        for x in range(17, 20):  # Displays stock of all bases
            name = str(i_sh["A" + str(x)].value)
            stock = str(i_sh["B" + str(x)].value)
            print(name + "=> " + stock + " in stock")
        i_wb.close()
    except Exception as e:
        print("Error occurred updating P and L statements.")
        writetoerrorlog("view_inventory", e)


def updatepandl(wb, total):  # updates P and L statements based on restock
    try:
        profit_sh = wb["P and L"]
        if total != 0:
            n = 2
            while True:  # Finds closest blank box and
                if profit_sh["D" + str(n)].value is None:  # inputs the total and customer
                    if profit_sh["C" + str(n - 1)].value == "Expenses":  # Determines the order number
                        order_num = 1
                    else:
                        ls = profit_sh["C" + str(n - 1)].value
                        order_num = int(ls[7]) + 1
                    profit_sh["C" + str(n)].value = f"Order #{order_num}"
                    profit_sh["D" + str(n)].value = -total  # Inputs order expense total
                    profit_sh["D" + str(n)].number_format = "$#,##0.00"  # Keeps money format in excel
                    break
                else:
                    n += 1
                    continue
    except Exception as e:
        print("Error occurred updating P and L statements.")
        writetoerrorlog("updatepandl", e)


def individual_restock():  # Restocks individual products based on user request
    try:
        i_wb = load_workbook("pizza.xlsx", data_only=False, keep_links=True)
        i_sh = i_wb["Inventory"]
        base = pd.read_csv('Prices/Base_raw.csv')  # Opens the files with pricing info
        toppings = pd.read_csv('Prices/Topping_raw.csv')
        base_diction = base.to_dict(orient='records')  # Converts files into lists of dictionaries
        toppings_diction = toppings.to_dict(orient='records')
        acceptable = []
        for x in chain(range(2, 15), range(17, 20)):
            name = str(i_sh["A" + str(x)].value).upper()
            acceptable.append(name)
        topping_total = 0
        size_total = 0
        while True:  # Asks user what they'd want to restock until finished
            stock = input("What would you like to restock? Type done when finished.")
            try:
                if stock.upper() == "DONE":  # Stops loop if user is finished
                    break
                elif stock.upper() in acceptable:  # Checks if response is in toppings or sizes list
                    for x in chain(range(2, 15), range(17, 20)):  # Fills
                        if stock.upper() == str(i_sh["A" + str(x)].value).upper():
                            if i_sh["B" + str(x)].value != 10:
                                ordered_amount = 10 - i_sh["B" + str(x)].value
                                i_sh["B" + str(x)].value = 10
                                if x in range(2, 15):
                                    top = toppings_diction[x - 2].get("Price") * ordered_amount
                                    topping_total += top
                                elif x in range(17, 20):
                                    size = base_diction[x - 17].get("Price") * ordered_amount
                                    size_total += size
                            else:
                                print("*Already at full stock*")
                else:
                    print("Invalid input.")
            except Exception as e:
                print("Error: Please input a valid number.")
                writetoerrorlog("individual_restock", e)
                continue
        total = size_total + topping_total
        updatepandl(i_wb, total)
        i_wb.save("pizza.xlsx")
        i_wb.close()
    except Exception as e:
        print("Error occurred updating Inventory.")
        writetoerrorlog("individual_restock", e)


def full_restock():  # Restocks all stock not at 10
    try:
        i_wb = load_workbook("pizza.xlsx", data_only=False, keep_links=True)
        i_sh = i_wb["Inventory"]
        base = pd.read_csv('Prices/Base_raw.csv')  # Opens the files with pricing info
        toppings = pd.read_csv('Prices/Topping_raw.csv')
        base_diction = base.to_dict(orient='records')  # Converts files into lists of dictionaries
        toppings_diction = toppings.to_dict(orient='records')
        topping_total = 0
        size_total = 0
        for x in chain(range(2, 15), range(17, 20)):  # Checks excel cells for low stock
            if i_sh["B" + str(x)].value < 10:
                ordered = 10 - i_sh["B" + str(x)].value
                i_sh["B" + str(x)].value = 10  # Fills stock
                if x in range(2, 15):
                    top = toppings_diction[x - 2].get("Price") * ordered  # Determines topping prices
                    topping_total += top
                elif x in range(17, 20):
                    size = base_diction[x - 17].get("Price") * ordered  # Determines size prices
                    size_total += size
        total = size_total + topping_total
        updatepandl(i_wb, total)
        i_wb.save("pizza.xlsx")
        i_wb.close()
    except Exception as e:
        print("Error occurred doing full restock.")
        writetoerrorlog("full_restock", e)


def restock():
    # Displays menu and either restocks inventory to full or restocks
    # individual products based on user input
    try:
        print(" ")
        print("Select restock menu option.")
        print("1. Individually restock")
        print("2. Full Restock")
        i = input("=>")
        if i == "1":
            individual_restock()
        elif i == "2":
            full_restock()
        else:
            print("Invalid input.")
    except Exception as e:
        print("Error occurred restocking.")
        writetoerrorlog("restock", e)


def view_pl():  # Displays either Sales, Expenses, or P and L total
    try:
        print(" ")
        print("Select a P and L menu option")
        print("1. View Sales")
        print("2. View Expenses")
        print("3. View P and L total")
        option = input("=>")
        if option == "1":
            profit_wb = load_workbook("pizza.xlsx", data_only=True, keep_links=True)
            profit_sh = profit_wb["P and L"]
            print(" ")
            print("Sales")
            x = 2
            while True:  # Displays sales of P and L sheet
                name = str(profit_sh["A" + str(x)].value)
                total = profit_sh["B" + str(x)].value
                if total is None:
                    break
                else:
                    print(name + "      $" + "{:.2f}".format(total))
                    x += 1
            wb, app = force_calc()                # Forces calculation so Nonetype is not displayed
            profit_sh = wb.sheets["P and L"]
            sales = "{:.2f}".format(profit_sh["F1"].value)
            print(" ")
            print(f"Sales Total: ${sales}")
            wb.close()
            app.quit()
        elif option == "2":
            profit_wb = load_workbook("pizza.xlsx", data_only=True, keep_links=True)
            profit_sh = profit_wb["P and L"]
            print(" ")
            print("Expenses")
            x = 2
            while True:  # Displays expenses of P and L sheet
                order_num = str(profit_sh["C" + str(x)].value)
                total = profit_sh["D" + str(x)].value
                if total is None:
                    break
                else:
                    print(order_num + "      $" + "{:.2f}".format(total))
                    x += 1
            wb, app = force_calc()  # Forces calculation so Nonetype is not displayed
            wb = xw.Book("pizza.xlsx")
            profit_sh = wb.sheets["P and L"]
            expenses = "{:.2f}".format(profit_sh["F2"].value)
            print(" ")
            print(f"Expenses Total: ${expenses}")
            wb.close()
            app.quit()
        elif option == "3":  # Displays overall P and L total
            wb, app = force_calc()      # Forces calculation so Nonetype is not displayed
            wb.app.calculate()
            profit_sh = wb.sheets["P and L"]
            prof = "{:.2f}".format(profit_sh["F4"].value)
            print(" ")
            print(f"P and L: ${prof}")
            wb.close()
            app.quit()
        else:
            print("Please input valid response.")

    except Exception as e:
        print("Error occurred displaying P and L statements.")
        writetoerrorlog("view_pl", e)


def balance():
    # Displays the account balance
    try:
        wb, app = force_calc()      # Forces calculation so Nonetype is not displayed
        profit_sh = wb.sheets["P and L"]
        account_balance = "{:.2f}".format(profit_sh["F6"].value)
        print(" ")
        print(f"Account balance: ${account_balance}")
        wb.close()
        app.quit()
    except Exception as e:
        print("Error occurred displaying balance.")
        writetoerrorlog("functions", e)


def force_calc():
    try:
        app = xw.App(visible=False)  # Forces calculation so Nonetype is not displayed
        wb = xw.Book("pizza.xlsx")
        wb.app.calculate()
        return wb, app
    except Exception as e:
        print("Error calculating values")
        writetoerrorlog("functions", e)



